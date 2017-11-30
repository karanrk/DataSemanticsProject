from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut
import csv
import openpyxl
from openpyxl import Workbook
import xlsxwriter


def get_latlon(area):
	geolocator = Nominatim()
	try:
		loc = geolocator.geocode(area)
	
		k=[0] * 2
		def try1(loc,area):
		
			if(area==None or len(area.split(" ")) <= 2):
				k[0],k[1]=0,0
				return k
			if(loc!=None):
				k[0],k[1]=loc.latitude,loc.longitude
				return k
			else:
				area=' '.join(area.split(' ')[1:])
				print "area splitting",area
				return get_latlon(area)
	except GeocoderTimedOut:
		return get_latlon(area)
		
	k=try1(loc,area)
	return k
			# k[0],k[1]=0,0
			# return k
			# loc=' '.join(loc.split(' ')[1:])
			
			
		
def getdata(file):
	
	#print filel
	book = openpyxl.load_workbook(file)
	sheet = book.active
	cells = sheet['B2':'F399']
	cells1 =sheet['A2':'A399']
	k=[]
	q=[]
	#print cells
	for c1, c2, c3 ,c4,c5  in cells:
		if(c2.value!=None):
			k.append('{} {} {} {} {}'.format(c1.value,c2.value,c3.value,c4.value,c5.value))
		else:
			k.append('{} {} {} {}'.format(c1.value,c3.value,c4.value,c5.value))
	#print k
	for c in cells1:
		#print c
		q=[sheet.cell(row=i,column=1).value for i in range(2,400)]

	l=1
	work = xlsxwriter.Workbook("//home/karan/Downloads/trial.xlsx")
	sheet1 = work.add_worksheet("Sheet 1")
	sheet1.write(0, 0, "Area")
	sheet1.write(0, 1, "Coordinates")
	sheet1.write(0,2,"Resource Name")
	for i in k:
		print i
		j=get_latlon(i)
		p="".join([str(x) for x in j])
		sheet1.write(l,0,i)
		sheet1.write(l,1,p)
		l+=1
	u=1
	for o in q:
		sheet1.write(u,2,o)
		u+=1
	#work.save()
	work.close()

#l=get_latlon("419 west amaryllis drive")
# print l[0]
# print l[1]

# s=get_latlon("PO Box 7083 Indianapolis IN 46207")
# print s[1]
getdata("/home/karan/Downloads/a.xlsx")
# l=getdata("/home/karan/Downloads/disability.xlsx")
# p={}
# for i in l:
	
# 		print i
# 		j=get_latlon(i)
# 		p[i]=j
	
		

